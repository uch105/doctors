import os
import sys
import django
from decouple import config

project_root = config("PROJECT_ROOT")
sys.path.append(project_root)

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'doctors.settings')
django.setup()

from dapp.models import Drug
from openpyxl import load_workbook

#path = 'C:\\Users\\uch\\Downloads\\projects\\doctors\\dapp\\drugs.xlsx'
path = '/var/www/doctors/dapp/drugs.xlsx'

workbook = load_workbook(path)
sheet = workbook.active

counter = 0

for row in range(1,sheet.max_row+1):
    manufacturer = sheet[f'A{row}'].value
    brand = sheet[f'I{row}'].value
    generic = sheet[f'C{row}'].value
    drugs_dose = sheet[f'D{row}'].value
    drugs_type = sheet[f'E{row}'].value
    price = sheet[f'F{row}'].value
    drugs_for = sheet[f'G{row}'].value
    drug_id = sheet[f'H{row}'].value

    Drug.objects.update_or_create(
        drug_id=drug_id,
        brand=brand,
        generic=generic,
        manufacturer=manufacturer,
        price=price,
        class_name = "No class",
        drugs_type = drugs_type,
        drugs_for = drugs_for,
        drugs_dose = drugs_dose,
    )

    counter+=1
    print(str(counter) + "Drugs imported successfully!")

    #print(str(row) + " drugs added")

print(str(counter) + "Drugs imported successfully!")